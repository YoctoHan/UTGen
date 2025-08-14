#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
通用工具函数模块
统一管理重复的工具函数，避免代码重复
增强了错误处理、日志记录和重试机制
"""
from typing import Union, List, Optional, Dict, Any
from pathlib import Path
import re
import os
import csv
import time
import json
import hashlib
from functools import wraps
from openai import OpenAI, RateLimitError
import logging
from datetime import datetime

# 配置日志
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)


# =============================================================================
# 装饰器和辅助函数
# =============================================================================

def retry_on_exception(max_retries: int = 3, delay: float = 1.0, backoff: float = 2.0):
    """
    重试装饰器，用于处理可能失败的操作
    
    Args:
        max_retries: 最大重试次数
        delay: 初始延迟时间（秒）
        backoff: 延迟时间的指数增长因子
    """
    def decorator(func):
        @wraps(func)
        def wrapper(*args, **kwargs):
            attempt = 0
            current_delay = delay
            
            while attempt < max_retries:
                try:
                    return func(*args, **kwargs)
                except Exception as e:
                    attempt += 1
                    if attempt >= max_retries:
                        logger.error(f"{func.__name__} 失败 {max_retries} 次后放弃: {str(e)}")
                        raise
                    
                    logger.warning(f"{func.__name__} 失败 (尝试 {attempt}/{max_retries}): {str(e)}")
                    logger.info(f"等待 {current_delay:.1f} 秒后重试...")
                    time.sleep(current_delay)
                    current_delay *= backoff
            
            return None
        return wrapper
    return decorator


def validate_path(path: Union[str, Path], must_exist: bool = False) -> Optional[Path]:
    """
    验证路径的有效性
    
    Args:
        path: 要验证的路径
        must_exist: 是否必须存在
    
    Returns:
        Path对象，如果无效则返回None
    """
    try:
        path_obj = Path(path) if not isinstance(path, Path) else path
        
        if must_exist and not path_obj.exists():
            logger.error(f"路径不存在: {path}")
            return None
            
        return path_obj
    except Exception as e:
        logger.error(f"路径验证失败 {path}: {str(e)}")
        return None


# =============================================================================
# 缓存管理
# =============================================================================

class CacheManager:
    """简单的文件缓存管理器"""
    
    def __init__(self, cache_dir: str = ".cache"):
        self.cache_dir = Path(cache_dir)
        self.cache_dir.mkdir(exist_ok=True)
    
    def get_cache_key(self, content: str) -> str:
        """生成缓存键"""
        return hashlib.md5(content.encode()).hexdigest()
    
    def get(self, key: str) -> Optional[str]:
        """获取缓存内容"""
        cache_file = self.cache_dir / f"{key}.json"
        if cache_file.exists():
            try:
                with open(cache_file, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                    # 检查缓存是否过期（24小时）
                    if time.time() - data.get('timestamp', 0) < 86400:
                        logger.info(f"使用缓存: {key}")
                        return data.get('content')
            except Exception as e:
                logger.warning(f"读取缓存失败: {str(e)}")
        return None
    
    def set(self, key: str, content: str):
        """设置缓存内容"""
        cache_file = self.cache_dir / f"{key}.json"
        try:
            with open(cache_file, 'w', encoding='utf-8') as f:
                json.dump({
                    'content': content,
                    'timestamp': time.time()
                }, f, ensure_ascii=False, indent=2)
            logger.info(f"已缓存: {key}")
        except Exception as e:
            logger.warning(f"写入缓存失败: {str(e)}")


# 全局缓存实例
cache_manager = CacheManager()


# =============================================================================
# 文件操作工具
# =============================================================================

def get_cpp_files(source_paths: List[Union[str, Path]], 
                  exclude_keywords: Optional[List[str]] = None,
                  max_depth: int = 10) -> List[Path]:
    """
    收集指定路径下的C++源码文件
    
    Args:
        source_paths: 源码路径列表
        exclude_keywords: 排除关键词列表
        max_depth: 最大递归深度
    
    Returns:
        list: C++文件路径列表
    """
    if exclude_keywords is None:
        exclude_keywords = ['test', 'benchmark', 'sample', 'example', '.git', 'build', 'cmake-build']
    
    cpp_extensions = {'.cpp', '.cc', '.cxx', '.c++', '.c', '.hpp', '.h', '.hxx', '.h++', '.hh'}
    cpp_files = []
    processed_paths = set()
    
    for path_str in source_paths:
        path = validate_path(path_str)
        if not path:
            continue
        
        # 避免处理重复路径
        abs_path = path.resolve()
        if abs_path in processed_paths:
            continue
        processed_paths.add(abs_path)
            
        if path.is_file():
            if path.suffix.lower() in cpp_extensions:
                cpp_files.append(path)
                logger.debug(f"添加文件: {path}")
        elif path.is_dir():
            # 使用递归深度限制避免无限递归
            for file_path in _walk_directory(path, max_depth):
                if (file_path.is_file() and 
                    file_path.suffix.lower() in cpp_extensions and
                    not any(keyword in str(file_path).lower() for keyword in exclude_keywords)):
                    cpp_files.append(file_path)
                    logger.debug(f"添加文件: {file_path}")
    
    unique_files = sorted(set(cpp_files))
    logger.info(f"共找到 {len(unique_files)} 个C++文件")
    return unique_files


def _walk_directory(directory: Path, max_depth: int, current_depth: int = 0):
    """递归遍历目录，带深度限制"""
    if current_depth >= max_depth:
        logger.warning(f"达到最大递归深度 {max_depth}，跳过子目录: {directory}")
        return
    
    try:
        for item in directory.iterdir():
            if item.is_file():
                yield item
            elif item.is_dir() and not item.name.startswith('.'):
                yield from _walk_directory(item, max_depth, current_depth + 1)
    except PermissionError:
        logger.warning(f"权限不足，无法访问: {directory}")
    except Exception as e:
        logger.error(f"遍历目录失败 {directory}: {str(e)}")


# 用于匹配文件最开头的 C/Java/JS 风格块注释（/** ... */ 或 /* ... */）
_COMMENT_BLOCK_RE = re.compile(r'^\s*/\*{1,2}.*?\*/', re.DOTALL)

def _strip_huawei_apache_header(text: str) -> str:
    """
    移除文件开头的华为 Apache 2.0 版权声明块注释。
    仅当首个块注释包含关键字时才删除，避免误删其他注释。
    """
    if not text:
        return text

    # 去除 UTF-8 BOM（如果存在）
    if text.startswith("\ufeff"):
        text = text.lstrip("\ufeff")

    m = _COMMENT_BLOCK_RE.match(text)
    if not m:
        return text

    block = m.group(0)
    lower = block.lower()

    # 关键字判断，确保是目标声明再移除
    has_apache = "licensed under the apache license" in lower and "version 2.0" in lower
    has_huawei = "huawei technologies" in lower

    if has_apache and has_huawei:
        # 移除块注释及其后紧随的空行
        stripped = text[m.end():]
        # 只清理开头多余空白行，不影响正文缩进
        stripped = stripped.lstrip("\r\n")
        return stripped

    return text

@retry_on_exception(max_retries=3)
def read_file_content(file_path: Union[str, Path], 
                     max_size: int = 2*1024*1024,
                     encodings: List[str] = None) -> str:
    """
    读取文件内容，支持多种编码和错误恢复；自动移除文件开头的华为 Apache 2.0 版权声明（块注释）。

    Args:
        file_path: 文件路径
        max_size: 文件最大大小限制（字节）
        encodings: 尝试的编码列表

    Returns:
        str: 文件内容，读取失败返回空字符串
    """
    if encodings is None:
        # 增加 utf-8-sig 以更好处理带 BOM 的文件
        encodings = ['utf-8', 'utf-8-sig', 'gbk', 'gb2312', 'iso-8859-1', 'ascii']
    
    path = validate_path(file_path, must_exist=True)
    if not path:
        return ""
    
    try:
        file_size = path.stat().st_size
        if file_size > max_size:
            logger.warning(f"文件过大，跳过: {path} ({file_size:,} bytes > {max_size:,} bytes)")
            return ""
        
        # 尝试不同的编码
        for encoding in encodings:
            try:
                with open(path, 'r', encoding=encoding) as f:
                    content = f.read()
                    content = _strip_huawei_apache_header(content)
                    logger.debug(f"成功读取文件 (编码: {encoding}): {path}")
                    return content
            except UnicodeDecodeError:
                continue
        
        # 如果所有编码都失败，尝试忽略错误
        with open(path, 'r', encoding='utf-8', errors='ignore') as f:
            content = f.read()
            content = _strip_huawei_apache_header(content)
            logger.warning(f"使用错误忽略模式读取文件: {path}")
            return content
            
    except Exception as e:
        logger.error(f"读取文件失败 {path}: {str(e)}")
        return ""


def save_file_content(content: str, file_path: Union[str, Path], 
                     backup: bool = False) -> bool:
    """
    保存内容到文件，支持备份
    
    Args:
        content: 要保存的内容
        file_path: 文件路径
        backup: 是否创建备份
    
    Returns:
        bool: 保存是否成功
    """
    path = validate_path(file_path)
    if not path:
        return False
    
    try:
        # 创建父目录
        path.parent.mkdir(parents=True, exist_ok=True)
        
        # 如果文件存在且需要备份
        if backup and path.exists():
            backup_path = path.with_suffix(f'.{datetime.now().strftime("%Y%m%d_%H%M%S")}.bak')
            path.rename(backup_path)
            logger.info(f"创建备份: {backup_path}")
        
        # 保存文件
        with open(path, 'w', encoding='utf-8') as f:
            f.write(content)
        
        logger.info(f"文件已保存: {path} ({len(content):,} 字符)")
        return True
        
    except Exception as e:
        logger.error(f"保存文件失败 {path}: {str(e)}")
        return False


# =============================================================================
# CSV处理工具
# =============================================================================

def read_csv_content(csv_file_path: Union[str, Path], 
                    delimiter: str = ',',
                    format_as_table: bool = True) -> str:
    """
    读取CSV文件内容并格式化显示
    
    Args:
        csv_file_path: CSV文件路径
        delimiter: 分隔符
        format_as_table: 是否格式化为表格
    
    Returns:
        str: 格式化后的CSV内容，读取失败返回空字符串
    """
    path = validate_path(csv_file_path, must_exist=True)
    if not path:
        return ""
    
    try:
        with open(path, 'r', encoding='utf-8', newline='') as f:
            reader = csv.reader(f, delimiter=delimiter)
            rows = list(reader)
        
        if not rows:
            logger.warning(f"CSV文件为空: {path}")
            return ""
        
        if format_as_table:
            # 格式化为Markdown表格
            lines = []
            for i, row in enumerate(rows):
                if i == 0:
                    # 标题行
                    lines.append("| " + " | ".join(row) + " |")
                    lines.append("|" + "|".join([" --- " for _ in row]) + "|")
                else:
                    # 数据行
                    lines.append("| " + " | ".join(row) + " |")
            
            result = "\n".join(lines)
        else:
            # 原始CSV格式
            result = "\n".join([delimiter.join(row) for row in rows])
        
        logger.info(f"成功读取CSV文件: {path} ({len(rows)} 行)")
        return result
    
    except Exception as e:
        logger.error(f"读取CSV文件失败 {path}: {str(e)}")
        return ""


def save_csv_content(csv_lines: List[str], output_file: Union[str, Path],
                    delimiter: str = ',') -> bool:
    """
    保存CSV内容到文件
    
    Args:
        csv_lines: CSV行列表
        output_file: 输出文件路径
        delimiter: 分隔符
    
    Returns:
        bool: 保存是否成功
    """
    if not csv_lines:
        logger.warning("没有有效的CSV内容")
        return False
    
    path = validate_path(output_file)
    if not path:
        return False
    
    try:
        path.parent.mkdir(parents=True, exist_ok=True)
        
        # 解析CSV行并重新格式化
        rows = []
        for line in csv_lines:
            if line.strip():
                # 尝试解析不同格式的CSV行
                if '|' in line:  # Markdown表格格式
                    row = [cell.strip() for cell in line.split('|') if cell.strip()]
                else:  # 标准CSV格式
                    row = [cell.strip() for cell in line.split(delimiter)]
                
                # 跳过分隔线
                if not all('---' in cell for cell in row):
                    rows.append(row)
        
        # 写入CSV文件
        with open(path, 'w', encoding='utf-8', newline='') as f:
            writer = csv.writer(f, delimiter=delimiter)
            writer.writerows(rows)
        
        logger.info(f"CSV文件已保存: {path} ({len(rows)} 行)")
        return True
        
    except Exception as e:
        logger.error(f"保存CSV文件失败 {path}: {str(e)}")
        return False


def save_xlsx_content(csv_lines: List[str], output_file: Union[str, Path]) -> bool:
    """
    保存CSV内容到Excel文件（XLSX格式）
    正确处理包含逗号的单元格内容，如列表格式的数据
    
    Args:
        csv_lines: CSV行列表
        output_file: 输出文件路径
    
    Returns:
        bool: 保存是否成功
    """
    if not csv_lines:
        logger.warning("没有有效的内容")
        return False
    
    path = validate_path(output_file)
    if not path:
        return False
    
    try:
        import openpyxl
        from openpyxl import Workbook
        import csv
        import io
        
        path.parent.mkdir(parents=True, exist_ok=True)
        
        # 创建新的工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = "TestParameters"
        
        # 解析CSV行
        for row_idx, line in enumerate(csv_lines, 1):
            if line.strip():
                # 使用csv模块正确解析包含逗号的内容
                try:
                    reader = csv.reader(io.StringIO(line))
                    row_data = next(reader)
                    
                    # 写入Excel
                    for col_idx, value in enumerate(row_data, 1):
                        # 清理值
                        value = value.strip()
                        
                        # 尝试转换数值类型，但保留列表格式的字符串
                        if value and not (value.startswith('[') or value.startswith('(') or value.startswith('{')):
                            try:
                                # 尝试转换为整数
                                if '.' not in value:
                                    cell_value = int(value)
                                else:
                                    # 尝试转换为浮点数
                                    cell_value = float(value)
                            except ValueError:
                                # 保持为字符串
                                cell_value = value
                        else:
                            # 保持列表/元组/字典格式的字符串
                            cell_value = value
                        
                        ws.cell(row=row_idx, column=col_idx, value=cell_value)
                        
                except Exception as e:
                    logger.warning(f"解析第{row_idx}行失败: {e}")
                    # 如果csv解析失败，尝试简单分割
                    row_data = [cell.strip() for cell in line.split(',')]
                    for col_idx, value in enumerate(row_data, 1):
                        ws.cell(row=row_idx, column=col_idx, value=value)
        
        # 自动调整列宽
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)  # 最大宽度限制为50
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # 保存文件
        wb.save(path)
        
        logger.info(f"Excel文件已保存: {path} ({len(csv_lines)} 行)")
        return True
        
    except ImportError:
        logger.error("需要安装openpyxl库: pip install openpyxl")
        return False
    except Exception as e:
        logger.error(f"保存Excel文件失败 {path}: {str(e)}")
        return False


# =============================================================================
# Excel处理工具
# =============================================================================

def read_excel_content(excel_path: Union[str, Path], 
                      sheet_name: Optional[str] = None) -> str:
    """
    读取Excel文件内容，转换为文本描述
    
    Args:
        excel_path: Excel文件路径
        sheet_name: 指定的工作表名称，None表示读取所有
    
    Returns:
        str: Excel内容的文本格式
    """
    path = validate_path(excel_path, must_exist=True)
    if not path:
        return ""
    
    try:
        import pandas as pd
        
        content = []
        
        if sheet_name:
            # 读取指定工作表
            df = pd.read_excel(path, sheet_name=sheet_name)
            content.append(f"Sheet: {sheet_name}")
            content.append(df.to_string(index=False))
        else:
            # 读取所有工作表
            excel_file = pd.ExcelFile(path)
            for sheet in excel_file.sheet_names:
                df = pd.read_excel(path, sheet_name=sheet)
                content.append(f"Sheet: {sheet}")
                content.append(df.to_string(index=False))
                content.append("")
        
        result = "\n".join(content)
        logger.info(f"成功读取Excel文件: {path}")
        return result
        
    except ImportError:
        logger.error("需要安装pandas和openpyxl库: pip install pandas openpyxl")
        return f"[无法读取Excel文件 {path}: 缺少pandas库]"
    except Exception as e:
        logger.error(f"读取Excel文件失败 {path}: {str(e)}")
        return f"[无法读取Excel文件 {path}: {e}]"


# =============================================================================
# 模型调用工具
# =============================================================================

class ModelCaller:
    """增强的模型调用器，支持缓存和更好的错误处理"""
    
    def __init__(self, api_key: str, base_url: str, model_name: str,
                 use_cache: bool = True):
        self.api_key = api_key
        self.base_url = base_url
        self.model_name = model_name
        self.use_cache = use_cache
        self.client = OpenAI(api_key=api_key, base_url=base_url)
    
    def call(self, prompt: str, system_message: Optional[str] = None,
             max_retries: int = 5, temperature: float = 0.7,
             max_tokens: int = 65536) -> str:
        """
        调用模型生成内容
        
        Args:
            prompt: 输入提示
            system_message: 系统消息
            max_retries: 最大重试次数
            temperature: 生成温度
            max_tokens: 最大token数
        
        Returns:
            str: 生成的内容
        """
        if system_message is None:
            system_message = "你是一个专业的C++程序员和测试工程师。请根据提供的代码为算子类写一个完整的单元测试(UT)。请直接输出UT代码，不需要额外的说明。"
        
        # 检查缓存
        if self.use_cache:
            cache_key = cache_manager.get_cache_key(f"{self.model_name}:{system_message}:{prompt}")
            cached_result = cache_manager.get(cache_key)
            if cached_result:
                return cached_result
        
        wait_time = 10
        attempt = 0
        
        while attempt < max_retries:
            attempt += 1
            
            try:
                logger.info(f"调用模型 {self.model_name} (尝试 {attempt}/{max_retries})")
                
                response = self.client.chat.completions.create(
                    model=self.model_name,
                    messages=[
                        {"role": "system", "content": system_message},
                        {"role": "user", "content": prompt}
                    ],
                    temperature=temperature,
                    max_tokens=max_tokens,
                    stream=True,
                    timeout=600
                )
                
                # 收集流式响应
                result = []
                for chunk in response:
                    if chunk.choices[0].delta.content:
                        result.append(chunk.choices[0].delta.content)
                
                full_result = "".join(result).strip()
                
                if full_result:
                    logger.info(f"模型调用成功，生成 {len(full_result):,} 字符")
                    
                    # 缓存结果
                    if self.use_cache:
                        cache_manager.set(cache_key, full_result)
                    
                    return full_result
                else:
                    logger.warning("模型返回空内容")
                    
            except RateLimitError as e:
                wait_time = min(wait_time * 2, 300)  # 最大等待5分钟
                logger.warning(f"达到速率限制: {str(e)}")
                logger.info(f"等待 {wait_time} 秒后重试...")
                time.sleep(wait_time)
                
            except Exception as e:
                logger.error(f"模型调用出错: {str(e)}")
                if attempt < max_retries:
                    logger.info(f"等待 10 秒后重试...")
                    time.sleep(10)
        
        logger.error(f"在 {max_retries} 次尝试后模型调用失败")
        return ""


def call_model(prompt: str, api_key: str, base_url: str, model_name: str,
              system_message: Optional[str] = None, max_retries: int = 5) -> str:
    """
    调用模型生成内容（向后兼容的接口）
    
    Args:
        prompt: 输入提示
        api_key: API密钥
        base_url: API基础URL
        model_name: 模型名称
        system_message: 系统消息
        max_retries: 最大重试次数
    
    Returns:
        str: 生成的内容
    """
    caller = ModelCaller(api_key, base_url, model_name)
    return caller.call(prompt, system_message, max_retries)


# =============================================================================
# 通用工具函数
# =============================================================================

def create_timestamped_dir(base_name: str, parent_dir: str = "runs") -> Path:
    """
    创建带时间戳的目录
    
    Args:
        base_name: 基础目录名
        parent_dir: 父目录
    
    Returns:
        Path: 创建的目录路径
    """
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    dir_path = Path(parent_dir) / f"{timestamp}_{base_name.lower()}"
    dir_path.mkdir(parents=True, exist_ok=True)
    logger.info(f"创建目录: {dir_path}")
    return dir_path


def log_message(message: str, log_file: Optional[Union[str, Path]] = None,
               level: str = "INFO"):
    """
    输出日志信息
    
    Args:
        message: 日志消息
        log_file: 日志文件路径
        level: 日志级别
    """
    # 输出到控制台
    log_func = getattr(logger, level.lower(), logger.info)
    log_func(message)
    
    # 输出到文件
    if log_file:
        try:
            path = validate_path(log_file)
            if path:
                with open(path, 'a', encoding='utf-8') as f:
                    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                    f.write(f"[{timestamp}] [{level}] {message}\n")
        except Exception as e:
            logger.error(f"写入日志文件失败: {str(e)}")


def validate_config(config: Dict[str, Any], required_keys: List[str]) -> bool:
    """
    验证配置字典
    
    Args:
        config: 配置字典
        required_keys: 必需的键列表
    
    Returns:
        bool: 配置是否有效
    """
    missing_keys = []
    
    for key in required_keys:
        if key not in config or not config[key]:
            missing_keys.append(key)
            logger.error(f"配置缺失或为空: {key}")
    
    if missing_keys:
        logger.error(f"配置验证失败，缺少以下配置: {', '.join(missing_keys)}")
        return False
    
    logger.info("配置验证通过")
    return True


def format_file_size(size_bytes: int) -> str:
    """
    格式化文件大小显示
    
    Args:
        size_bytes: 字节数
    
    Returns:
        str: 格式化后的大小字符串
    """
    for unit in ['B', 'KB', 'MB', 'GB']:
        if size_bytes < 1024.0:
            return f"{size_bytes:.2f} {unit}"
        size_bytes /= 1024.0
    return f"{size_bytes:.2f} TB"


def get_file_stats(file_path: Union[str, Path]) -> Dict[str, Any]:
    """
    获取文件统计信息
    
    Args:
        file_path: 文件路径
    
    Returns:
        dict: 文件统计信息
    """
    path = validate_path(file_path, must_exist=True)
    if not path:
        return {}
    
    try:
        stat = path.stat()
        return {
            'size': stat.st_size,
            'size_formatted': format_file_size(stat.st_size),
            'modified': datetime.fromtimestamp(stat.st_mtime),
            'created': datetime.fromtimestamp(stat.st_ctime),
            'lines': len(path.read_text(encoding='utf-8').splitlines()) if path.suffix in ['.py', '.cpp', '.h', '.txt'] else None
        }
    except Exception as e:
        logger.error(f"获取文件统计信息失败 {path}: {str(e)}")
        return {}


# =============================================================================
# 初始化和清理
# =============================================================================

def cleanup_old_cache(days: int = 7):
    """清理旧的缓存文件"""
    cache_dir = Path(".cache")
    if not cache_dir.exists():
        return
    
    current_time = time.time()
    for cache_file in cache_dir.glob("*.json"):
        try:
            if current_time - cache_file.stat().st_mtime > days * 86400:
                cache_file.unlink()
                logger.info(f"删除旧缓存: {cache_file.name}")
        except Exception as e:
            logger.warning(f"清理缓存失败: {str(e)}")


# 启动时清理旧缓存
cleanup_old_cache()